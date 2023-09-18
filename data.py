from tkinter import*
import openpyxl  

class Register:
    def __init__(self,root):
        self.root = root;
        self.root.title("Registration Form")
        self.root.geometry("500x500+500+100")
        
        title = Label(root,text="USER REGISTRATION ",font=("times new roman",20,"bold")).place(x=100,y=35);

        self.name_var = StringVar() 
        name = Label(root,text="Name :",font=("times new roman",15)).place(x=60,y=100);
        name_en = Entry(root,font=("times new roman",15), textvariable=self.name_var).place(x=140,y=100,width=200);

        self.email_var = StringVar()
        email = Label(root,text="Email :",font=("times new roman",15)).place(x=60,y=150);
        email_en = Entry(root,font=("times new roman",15),textvariable=self.email_var).place(x=140,y=150,width=200);

        self.contact_var = StringVar()
        contact_no = Label(root,text="Contact No :",font=("times new roman",15)).place(x=30,y=200);
        contact_en = Entry(root,font=("times new roman",15),textvariable=self.contact_var).place(x=140,y=200,width=200);

        self.address_var = StringVar()
        address = Label(root,text="Address :",font=("times new roman",15)).place(x=50,y=250);
        address_en = Entry(root,font=("times new roman",15),textvariable=self.address_var).place(x=140,y=250,width=200);

        btn = Button(root,text="Submit",font=("times new roman",15) ,border=2,command=self.register_data).place(x=180,y=300 ,width=120);

    def register_data(self):
        name = self.name_var.get()
        email = self.email_var.get()
        contact = self.contact_var.get()
        address = self.address_var.get()

        workbook = openpyxl.Workbook()

        sheet = workbook.active

        try:
            workbook = openpyxl.load_workbook("myDataSheet.xlsx")
            sheet = workbook.active
        except FileNotFoundError:
            
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.cell(row=1, column=1, value="Name")
            sheet.cell(row=1, column=2, value="Email")
            sheet.cell(row=1, column=3, value="Contact No")
            sheet.cell(row=1, column=4, value="Address")

            if sheet.cell(row=1, column=1).value is None:
                sheet.cell(row=1, column=1, value="Name")
                sheet.cell(row=1, column=2, value="Email")
                sheet.cell(row=1, column=3, value="Contact No")
                sheet.cell(row=1, column=4, value="Address")

        row = sheet.max_row + 1
        sheet.cell(row=row, column=1, value=name)
        sheet.cell(row=row, column=2, value=email)
        sheet.cell(row=row, column=3, value=contact)
        sheet.cell(row=row, column=4, value=address)

        workbook.save("myDataSheet.xlsx")
        
        print("Data saved to myDataSheet.xlsx")

root = Tk()
obj= Register(root)
root.mainloop()